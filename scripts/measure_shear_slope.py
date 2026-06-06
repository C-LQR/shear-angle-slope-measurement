from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

import numpy as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ZH = {
    "fit_result": "\u62df\u5408\u7ed3\u679c",
    "image_name": "\u56fe\u50cf\u540d\u79f0",
    "region": "\u533a\u57df",
    "equation_header": "\u62df\u5408\u65b9\u7a0b\uff08\u5c40\u90e8\u5750\u6807\uff0cy\u8f74\u5411\u4e0a\uff09",
    "shear_angle": "\u526a\u5207\u89d2",
    "degree": "\u5ea6",
    "r2": "\u51b3\u5b9a\u7cfb\u6570",
    "n": "\u62df\u5408\u70b9\u6570 n",
    "roi": "ROI\uff08\u611f\u5174\u8da3\u533a\u57df\uff09\u5750\u6807",
    "reference_length": "\u76d2\u4f53\u6807\u5b9a\u957f\u5ea6(mm)",
    "reference_pixels": "\u6807\u5b9a\u50cf\u7d20\u957f\u5ea6(px)",
    "runout_pixels": "\u503e\u6cfb\u8ddd\u79bb(px)",
    "runout_distance": "\u503e\u6cfb\u8ddd\u79bb(mm)",
    "runout_label": "\u503e\u6cfb\u8ddd\u79bb",
    "group_id": "\u7ec4\u53f7",
    "shear_sheet": "\u526a\u5207\u89d2",
    "runout_sheet": "\u503e\u6cfb\u8ddd\u79bb",
    "material_slope": "\u76d2\u4f53\u5916\u6d41\u51fa\u7269\u6599\u659c\u5761",
    "former": "\u524d\u7aef",
    "latter": "\u540e\u7aef",
    "total": "\u6574\u4f53",
    "former_file": "\u524d\u8005",
    "latter_file": "\u540e\u8005",
    "total_file": "\u6574\u4f53",
    "fit_equation": "\u62df\u5408\u65b9\u7a0b",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Measure shear slope from a cropped particle-flow image.")
    parser.add_argument("--image", required=True, help="Source image path.")
    parser.add_argument("--root", default=None, help="Output root. Defaults to the source image parent.")
    parser.add_argument("--crop", nargs=4, type=int, metavar=("X0", "Y0", "X1", "Y1"), required=True)
    parser.add_argument("--fit-x-min", type=int, required=True, help="Original-image x minimum for fitting.")
    parser.add_argument("--fit-x-max", type=int, required=True, help="Original-image x maximum for fitting.")
    parser.add_argument("--fit-y-max", type=int, required=True, help="Original-image y maximum for top-surface fitting.")
    parser.add_argument("--threshold", type=float, default=95.0, help="Gray threshold for dark material segmentation.")
    parser.add_argument("--residual-threshold", type=float, default=10.0, help="Inlier residual threshold in pixels.")
    parser.add_argument("--workbook-name", default=None, help="Workbook name. Defaults to <root folder name>.xlsx.")
    parser.add_argument("--segment-mode", choices=["single", "auto"], default="single")
    parser.add_argument("--split-x", type=int, default=None, help="Manual split x for auto segment mode.")
    parser.add_argument("--flat-latter-angle-max", type=float, default=2.0, help="If auto-mode latter angle is at or below this value, keep only the former/front fit.")
    parser.add_argument("--reference-length-mm", type=float, default=100.0, help="Actual transparent box length used for pixel-to-mm scale.")
    return parser.parse_args()


def chinese_font(size: int) -> ImageFont.FreeTypeFont:
    for candidate in [r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\simsun.ttc"]:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    raise FileNotFoundError("No Chinese-capable Windows font found.")


def extract_top_envelope(rgb: np.ndarray, crop: tuple[int, int, int, int], threshold: float) -> np.ndarray:
    x0, y0, x1, y1 = crop
    roi = rgb[y0:y1, x0:x1]
    mask = roi.mean(axis=2) < threshold
    points: list[tuple[int, int]] = []
    for xi in range(mask.shape[1]):
        ys = np.where(mask[:, xi])[0]
        if len(ys) >= 5:
            top = int(np.median(ys[: min(5, len(ys))]))
            points.append((x0 + xi, y0 + top))
    if not points:
        raise RuntimeError("No material envelope points found. Check ROI and threshold.")
    return np.array(points)


def robust_line_fit(points: np.ndarray, residual_threshold: float) -> tuple[float, float, np.ndarray, float, float]:
    if len(points) < 2:
        raise RuntimeError("Too few points for line fitting.")
    a, b = np.linalg.lstsq(np.vstack([points[:, 0], np.ones(len(points))]).T, points[:, 1], rcond=None)[0]
    inliers = np.ones(len(points), dtype=bool)
    for _ in range(5):
        residual = np.abs(points[:, 1] - (a * points[:, 0] + b))
        inliers = residual < residual_threshold
        if inliers.sum() < 2:
            raise RuntimeError("Too few inliers for line fitting.")
        a, b = np.linalg.lstsq(np.vstack([points[inliers, 0], np.ones(inliers.sum())]).T, points[inliers, 1], rcond=None)[0]
    fit_points = points[inliers]
    y_pred = a * fit_points[:, 0] + b
    sse = float(np.sum((fit_points[:, 1] - y_pred) ** 2))
    sst = float(np.sum((fit_points[:, 1] - np.mean(fit_points[:, 1])) ** 2))
    r2 = 1 - sse / sst if sst else float("nan")
    return float(a), float(b), inliers, sse, r2


def ordinary_line_fit(points: np.ndarray) -> tuple[float, float, np.ndarray, float, float]:
    if len(points) < 2:
        raise RuntimeError("Too few points for line fitting.")
    a, b = np.linalg.lstsq(np.vstack([points[:, 0], np.ones(len(points))]).T, points[:, 1], rcond=None)[0]
    inliers = np.ones(len(points), dtype=bool)
    y_pred = a * points[:, 0] + b
    sse = float(np.sum((points[:, 1] - y_pred) ** 2))
    sst = float(np.sum((points[:, 1] - np.mean(points[:, 1])) ** 2))
    r2 = 1 - sse / sst if sst else float("nan")
    return float(a), float(b), inliers, sse, r2


def make_fit(name: str, label: str, points: np.ndarray, crop: tuple[int, int, int, int], residual_threshold: float, robust: bool) -> dict:
    if robust:
        a_down, b_down, inliers, sse, r2 = robust_line_fit(points, residual_threshold)
    else:
        a_down, b_down, inliers, sse, r2 = ordinary_line_fit(points)
    fit_points = points[inliers]
    x0, y0, _x1, y1 = crop
    crop_height = y1 - y0
    local_b_down = a_down * x0 + b_down - y0
    return {
        "name": name,
        "label": label,
        "points": points,
        "fit_points": fit_points,
        "a_down": a_down,
        "b_down": b_down,
        "slope_y_up": -a_down,
        "intercept_y_up": crop_height - local_b_down,
        "shear_angle": math.degrees(math.atan(abs(a_down))),
        "r2": r2,
        "sse": sse,
        "n": int(inliers.sum()),
        "x_range": (int(fit_points[:, 0].min()), int(fit_points[:, 0].max())),
    }


def choose_split(points: np.ndarray, residual_threshold: float) -> int:
    x_min, x_max = int(points[:, 0].min()), int(points[:, 0].max())
    min_points = max(40, min(120, len(points) // 5))
    best: tuple[float, int] | None = None
    for split in range(x_min + 80, x_max - 80 + 1, 10):
        left = points[points[:, 0] < split]
        right = points[points[:, 0] >= split]
        if len(left) < min_points or len(right) < min_points:
            continue
        try:
            _a1, _b1, _s1, sse1, _r21 = ordinary_line_fit(left)
            _a2, _b2, _s2, sse2, _r22 = ordinary_line_fit(right)
        except RuntimeError:
            continue
        score = sse1 + sse2
        if best is None or score < best[0]:
            best = (score, split)
    if best is None:
        raise RuntimeError("Could not choose a segment split. Check fit bounds.")
    return best[1]


def measure_runout_distance(
    rgb: np.ndarray,
    crop: tuple[int, int, int, int],
    outside: np.ndarray,
    threshold: float,
    reference_length_mm: float,
) -> dict:
    x0, y0, x1, y1 = crop
    roi = rgb[y0:y1, x0:x1]
    crop_h = y1 - y0

    dark_mask = roi.mean(axis=2) < threshold
    dark_counts = dark_mask.sum(axis=0)
    material_min_count = max(20, int(crop_h * 0.05))
    material_cols = np.where(dark_counts >= material_min_count)[0]
    if len(material_cols) == 0:
        raise RuntimeError("Could not find left material boundary for distance measurement.")
    material_left_x = x0 + int(material_cols.min())

    rgb_range = roi.max(axis=2) - roi.min(axis=2)
    gray = roi.mean(axis=2)
    upper_h = max(20, int(crop_h * 0.75))
    box_mask = (gray[:upper_h] >= 150) & (gray[:upper_h] <= 235) & (rgb_range[:upper_h] <= 8)
    box_counts = box_mask.sum(axis=0)
    box_cols = np.where(box_counts >= int(upper_h * 0.45))[0]
    if len(box_cols) == 0:
        raise RuntimeError("Could not find transparent box right boundary for distance measurement.")
    box_right_x = x0 + int(box_cols.max())

    rear_endpoint_x = int(outside[:, 0].max())
    reference_pixels = box_right_x - material_left_x
    if reference_pixels <= 0:
        raise RuntimeError("Invalid reference length in pixels.")
    runout_pixels = rear_endpoint_x - material_left_x
    mm_per_pixel = reference_length_mm / reference_pixels
    return {
        "material_left_x": material_left_x,
        "box_right_x": box_right_x,
        "rear_endpoint_x": rear_endpoint_x,
        "reference_length_mm": reference_length_mm,
        "reference_pixels": reference_pixels,
        "mm_per_pixel": mm_per_pixel,
        "runout_pixels": runout_pixels,
        "runout_distance_mm": runout_pixels * mm_per_pixel,
    }


def compute_results(
    image_path: Path,
    crop: tuple[int, int, int, int],
    fit_x_min: int,
    fit_x_max: int,
    fit_y_max: int,
    threshold: float,
    residual_threshold: float,
    segment_mode: str,
    split_x: int | None,
    flat_latter_angle_max: float,
    reference_length_mm: float,
) -> dict:
    image = Image.open(image_path).convert("RGBA")
    rgb = np.array(image.convert("RGB"))
    points = extract_top_envelope(rgb, crop, threshold)
    outside = points[(points[:, 0] >= fit_x_min) & (points[:, 0] < fit_x_max) & (points[:, 1] < fit_y_max)]
    if len(outside) < 2:
        raise RuntimeError("Too few fit points. Check fit bounds.")
    distance = measure_runout_distance(rgb, crop, outside, threshold, reference_length_mm)

    if segment_mode == "auto":
        split = split_x if split_x is not None else choose_split(outside, residual_threshold)
        former_points = outside[outside[:, 0] < split]
        latter_points = outside[outside[:, 0] >= split]
        former_fit = make_fit("former", ZH["former"], former_points, crop, residual_threshold, robust=False)
        latter_fit = make_fit("latter", ZH["latter"], latter_points, crop, residual_threshold, robust=False)
        if latter_fit["shear_angle"] <= flat_latter_angle_max:
            fits = [former_fit]
        else:
            fits = [
                former_fit,
                latter_fit,
                make_fit("total", ZH["total"], outside, crop, residual_threshold, robust=False),
            ]
    else:
        split = None
        fits = [make_fit("total", ZH["total"], outside, crop, residual_threshold, robust=True)]
    return {"image": image, "crop": crop, "fits": fits, "split_x": split, "distance": distance}


def draw_fit_image(base_crop: Image.Image, fit: dict, crop: tuple[int, int, int, int], fit_path: Path) -> None:
    x0, y0, _x1, _y1 = crop
    crop_img = base_crop.copy()
    draw = ImageDraw.Draw(crop_img, "RGBA")
    fit_points = fit["fit_points"]
    a_down = fit["a_down"]
    b_down = fit["b_down"]
    fit_x_min, fit_x_max = fit["x_range"]
    line_start = (fit_x_min - x0, int(round(a_down * fit_x_min + b_down)) - y0)
    line_end = (fit_x_max - x0, int(round(a_down * fit_x_max + b_down)) - y0)
    for px, py in fit_points[::5]:
        lx, ly = int(px - x0), int(py - y0)
        draw.ellipse((lx - 2, ly - 2, lx + 2, ly + 2), fill=(255, 210, 0, 140))
    draw.line([line_start, line_end], fill=(220, 20, 60, 255), width=5)
    for lx, ly in (line_start, line_end):
        draw.ellipse((lx - 5, ly - 5, lx + 5, ly + 5), fill=(220, 20, 60, 255), outline=(255, 255, 255, 255), width=2)

    font_text = chinese_font(18)
    font_em = chinese_font(20)
    if "distance" in fit:
        distance = fit["distance"]
        sx = int(distance["material_left_x"] - x0)
        ex = int(distance["rear_endpoint_x"] - x0)
        y_line = int(round(fit["a_down"] * distance["rear_endpoint_x"] + fit["b_down"])) - y0
        y_line = max(24, min(crop_img.height - 28, y_line - 28))
        if ex < sx:
            sx, ex = ex, sx
        dash = 14
        gap = 8
        cursor = sx
        while cursor < ex:
            draw.line([(cursor, y_line), (min(cursor + dash, ex), y_line)], fill=(220, 20, 60, 255), width=3)
            cursor += dash + gap
        for px in (sx, ex):
            draw.ellipse((px - 5, y_line - 5, px + 5, y_line + 5), fill=(220, 20, 60, 255), outline=(255, 255, 255, 255), width=2)
        label = f"{ZH['runout_label']} = {distance['runout_distance_mm']:.3f} mm"
        bbox = draw.textbbox((0, 0), label, font=font_text)
        label_x = max(6, min((sx + ex) // 2 - (bbox[2] - bbox[0]) // 2, crop_img.width - (bbox[2] - bbox[0]) - 12))
        label_y = max(6, y_line - 30)
        draw.rectangle((label_x - 6, label_y - 4, label_x + (bbox[2] - bbox[0]) + 6, label_y + (bbox[3] - bbox[1]) + 4), fill=(255, 255, 255, 220))
        draw.text((label_x, label_y), label, font=font_text, fill=(220, 20, 60, 255))
    lines = [
        f"{fit['label']}{ZH['fit_equation']}: y = {fit['slope_y_up']:.6f}x + {fit['intercept_y_up']:.6f}",
        f"{ZH['shear_angle']} = {fit['shear_angle']:.3f} deg ({ZH['degree']})",
        f"R^2 ({ZH['r2']}) = {fit['r2']:.4f}",
        f"{ZH['n']} = {fit['n']}",
    ]
    pad, line_h = 12, 27
    box_w = 440
    box_h = pad * 2 + line_h * len(lines)
    box_x = crop_img.width - box_w - 16
    box_y = 16
    draw.rounded_rectangle((box_x, box_y, box_x + box_w, box_y + box_h), radius=8, fill=(255, 255, 255, 230), outline=(80, 80, 80, 180), width=1)
    for i, text in enumerate(lines):
        draw.text((box_x + pad, box_y + pad + i * line_h), text, font=font_em if i == 0 else font_text, fill=(180, 20, 45, 255) if i == 0 else (30, 30, 30, 255))
    crop_img.save(fit_path, "PNG")


def save_images(results: dict, image_path: Path, root: Path) -> tuple[Path, list[Path]]:
    x0, y0, x1, y1 = results["crop"]
    out_dir = root / "Local"
    out_dir.mkdir(exist_ok=True)
    local_path = out_dir / f"{image_path.stem}_local.png"
    base_crop = results["image"].crop((x0, y0, x1, y1))
    base_crop.save(local_path, "PNG")

    prefix_by_name = {"former": ZH["former_file"], "latter": ZH["latter_file"], "total": ZH["total_file"]}
    fit_paths: list[Path] = []
    for fit in results["fits"]:
        prefix = prefix_by_name.get(fit["name"], fit["label"])
        fit_path = out_dir / f"{prefix}{image_path.stem}fit.png"
        fit["distance"] = results["distance"]
        draw_fit_image(base_crop, fit, results["crop"], fit_path)
        fit_paths.append(fit_path)
    return local_path, fit_paths


def style_table(ws, max_row: int, max_col: int, widths: list[int], number_formats: dict[int, str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            elif cell.column in number_formats:
                cell.number_format = number_formats[cell.column]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def result_sheet_sort_key(sheet_name: str) -> tuple[str, int, str]:
    stem = sheet_name[: -len(ZH["fit_result"])]
    match = re.fullmatch(r"([A-Za-z_]+)(\d+)", stem)
    if match:
        return (match.group(1).lower(), int(match.group(2)), stem)
    return (stem.lower(), 0, stem)


def group_id_from_image_name(image_name: str) -> str:
    return Path(str(image_name)).stem


def is_total_region(region: str) -> bool:
    return str(region).endswith(f"-{ZH['total']}")


def merge_group_cells(ws, group_col: int = 1, start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return
    merge_start = start_row
    previous = ws.cell(row=start_row, column=group_col).value
    for row_index in range(start_row + 1, ws.max_row + 2):
        current = ws.cell(row=row_index, column=group_col).value if row_index <= ws.max_row else None
        if current != previous:
            if previous is not None and row_index - merge_start > 1:
                ws.merge_cells(start_row=merge_start, start_column=group_col, end_row=row_index - 1, end_column=group_col)
            merge_start = row_index
            previous = current


def rebuild_summary_sheets(wb) -> None:
    result_names = [
        name
        for name in wb.sheetnames
        if name.endswith(ZH["fit_result"])
        and name not in {ZH["shear_sheet"], ZH["runout_sheet"]}
    ]
    result_names.sort(key=result_sheet_sort_key)

    summary_specs = [
        (
            ZH["shear_sheet"],
            [
                ZH["group_id"],
                ZH["region"],
                ZH["equation_header"],
                f"{ZH['shear_angle']} deg({ZH['degree']})",
                f"R^2({ZH['r2']})",
                ZH["n"],
                ZH["roi"],
            ],
            lambda row: [group_id_from_image_name(row[0]), row[1], row[2], row[3], row[4], row[5], row[6]],
            [16, 28, 34, 18, 18, 14, 28],
            {4: "0.000000", 5: "0.000000", 6: "0"},
            lambda _row: True,
            True,
        ),
        (
            ZH["runout_sheet"],
            [
                ZH["group_id"],
                ZH["region"],
                ZH["reference_length"],
                ZH["reference_pixels"],
                ZH["runout_pixels"],
                ZH["runout_distance"],
                ZH["roi"],
            ],
            lambda row: [group_id_from_image_name(row[0]), row[1], row[7], row[8], row[9], row[10], row[6]],
            [16, 28, 20, 20, 18, 18, 28],
            {3: "0.000", 4: "0", 5: "0", 6: "0.000"},
            lambda row: is_total_region(row[1]),
            False,
        ),
    ]

    for sheet_name, headers, row_builder, widths, number_formats, row_filter, merge_groups in summary_specs:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        ws.delete_rows(1, ws.max_row)
        rows = [headers]
        for result_name in result_names:
            result_ws = wb[result_name]
            for values in result_ws.iter_rows(min_row=2, max_col=11, values_only=True):
                if values[0] and row_filter(values):
                    rows.append(row_builder(values))
        for row_index, row_values in enumerate(rows, start=1):
            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
        style_table(ws, len(rows), len(headers), widths, number_formats)
        if merge_groups:
            merge_group_cells(ws)


def save_workbook(results: dict, image_path: Path, root: Path, workbook_name: str | None) -> Path:
    workbook_path = root / (workbook_name or f"{root.name}.xlsx")
    sheet_name = f"{image_path.stem}{ZH['fit_result']}"
    if workbook_path.exists():
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = sheet_name
    headers = [
        ZH["image_name"],
        ZH["region"],
        ZH["equation_header"],
        f"{ZH['shear_angle']} deg({ZH['degree']})",
        f"R^2({ZH['r2']})",
        ZH["n"],
        ZH["roi"],
        ZH["reference_length"],
        ZH["reference_pixels"],
        ZH["runout_pixels"],
        ZH["runout_distance"],
    ]
    rows = [headers]
    distance = results["distance"]
    for fit in results["fits"]:
        rows.append([
            image_path.name,
            f"{ZH['material_slope']}-{fit['label']}",
            f"y = {fit['slope_y_up']:.6f}x + {fit['intercept_y_up']:.6f}",
            fit["shear_angle"],
            fit["r2"],
            fit["n"],
            f"({','.join(str(v) for v in results['crop'])})",
            distance["reference_length_mm"],
            distance["reference_pixels"],
            distance["runout_pixels"],
            distance["runout_distance_mm"],
        ])
    for row_index, row_values in enumerate(rows, start=1):
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    style_table(
        ws,
        len(rows),
        11,
        [18, 28, 34, 18, 18, 14, 28, 20, 20, 18, 18],
        {4: "0.000000", 5: "0.000000", 6: "0", 8: "0.000", 9: "0", 10: "0", 11: "0.000"},
    )
    rebuild_summary_sheets(wb)
    wb.save(workbook_path)
    return workbook_path


def main() -> None:
    args = parse_args()
    image_path = Path(args.image)
    root = Path(args.root) if args.root else image_path.parent
    crop = tuple(args.crop)
    results = compute_results(
        image_path=image_path,
        crop=crop,
        fit_x_min=args.fit_x_min,
        fit_x_max=args.fit_x_max,
        fit_y_max=args.fit_y_max,
        threshold=args.threshold,
        residual_threshold=args.residual_threshold,
        segment_mode=args.segment_mode,
        split_x=args.split_x,
        flat_latter_angle_max=args.flat_latter_angle_max,
        reference_length_mm=args.reference_length_mm,
    )
    local_path, fit_paths = save_images(results, image_path, root)
    workbook_path = save_workbook(results, image_path, root, args.workbook_name)
    print(f"local_image={local_path}")
    for fit_path in fit_paths:
        print(f"annotated_image={fit_path}")
    print(f"workbook={workbook_path}")
    if results["split_x"] is not None:
        print(f"split_x={results['split_x']}")
    for fit in results["fits"]:
        print(
            f"{fit['name']}: equation=y = {fit['slope_y_up']:.6f}x + {fit['intercept_y_up']:.6f}; "
            f"shear_angle={fit['shear_angle']:.6f}; r2={fit['r2']:.6f}; n={fit['n']}"
        )
    distance = results["distance"]
    print(
        f"distance: reference_px={distance['reference_pixels']}; runout_px={distance['runout_pixels']}; "
        f"runout_mm={distance['runout_distance_mm']:.6f}"
    )


if __name__ == "__main__":
    main()
